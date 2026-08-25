import frappe
import openpyxl
import io
import json
from collections import defaultdict
from frappe.model.document import Document
from frappe.utils import flt, cint, getdate


class AccountOpeningOperations(Document):

    def autoname(self):
        # Format: 1038-June-2026
        if self.sol_id and self.month and self.year:
            self.name = f"{self.sol_id}-{self.month.capitalize()}-{self.year}"

    def validate(self):
        self.group_and_merge_zero_ip_funding()
        self.sort_child_tables_by_date()
        self.calculate_totals_and_percentages()

    # ------------------------------------------------------------------
    # ZERO IP FUNDING MERGE & SCHEME COUNT FORMATTING
    # ------------------------------------------------------------------
    def group_and_merge_zero_ip_funding(self):
        """
        Merges duplicate dates in table_zero_ip_funding into a single row.
        Calculates per-scheme counts and generates scheme_code_options as a JSON string
        with formatted schemes like "1003(3)" sorted by highest count first.
        """
        if not self.get("table_zero_ip_funding"):
            return

        # Record existing manual selections before merging
        existing_selections = {}
        for row in self.get("table_zero_ip_funding"):
            if row.ac_opening_date and row.scheme_code:
                d_str = str(getdate(row.ac_opening_date))
                if d_str not in existing_selections:
                    existing_selections[d_str] = str(row.scheme_code).strip()

        # Structure: date_str -> { scheme_code -> count }
        grouped_data = defaultdict(lambda: defaultdict(int))

        for row in self.get("table_zero_ip_funding"):
            if not row.ac_opening_date:
                continue

            date_str = str(getdate(row.ac_opening_date))
            scheme = str(row.scheme_code or "N/A").strip()

            # Clean raw scheme if already has parenthetical count
            if "(" in scheme:
                scheme = scheme.split("(")[0].strip()

            count = cint(row.zero_ip_funding) or 1
            grouped_data[date_str][scheme] += count

        self.set("table_zero_ip_funding", [])

        # Re-build table with deduplicated date rows
        for date_val, schemes_dict in grouped_data.items():
            total_date_count = sum(schemes_dict.values())

            # Sort schemes by count descending: e.g., [("1003", 3), ("1001", 2), ("1002", 1)]
            sorted_schemes = sorted(
                schemes_dict.items(), key=lambda x: x[1], reverse=True
            )

            # Format scheme list: ["1003(3)", "1001(2)", "1002(1)"]
            formatted_schemes = [f"{sch}({cnt})" for sch, cnt in sorted_schemes]
            options_json = json.dumps(formatted_schemes)

            prev_selected = existing_selections.get(date_val)
            chosen_scheme = formatted_schemes[0] if formatted_schemes else ""
            if prev_selected and prev_selected in formatted_schemes:
                chosen_scheme = prev_selected

            self.append(
                "table_zero_ip_funding",
                {
                    "ac_opening_date": date_val,
                    "scheme_code": chosen_scheme,
                    "scheme_code_options": options_json,
                    "zero_ip_funding": total_date_count,
                },
            )

    # ------------------------------------------------------------------
    # DATE SORTING: DESCENDING ORDER (Latest Date on top)
    # ------------------------------------------------------------------
    def sort_child_tables_by_date(self):
        # Sort table_dllf
        if self.get("table_dllf"):
            sorted_rows = sorted(
                self.get("table_dllf"),
                key=lambda x: (
                    getdate(str(x.date)) if x.date else getdate("1900-01-01")
                ),
                reverse=True,
            )
            self.set("table_dllf", [])
            for idx, row in enumerate(sorted_rows, start=1):
                row.idx = idx
                self.append("table_dllf", row)

        # Sort table_zero_ip_funding
        if self.get("table_zero_ip_funding"):
            sorted_rows_zero_ip = sorted(
                self.get("table_zero_ip_funding"),
                key=lambda x: (
                    getdate(str(x.ac_opening_date))
                    if x.ac_opening_date
                    else getdate("1900-01-01")
                ),
                reverse=True,
            )
            self.set("table_zero_ip_funding", [])
            for idx, row in enumerate(sorted_rows_zero_ip, start=1):
                row.idx = idx
                self.append("table_zero_ip_funding", row)

    def calculate_totals_and_percentages(self):
        total_ftr = 0
        total_ftnr = 0

        # Child table: table_dllf
        for row in self.get("table_dllf", []):
            total_ftr += cint(row.ftr)
            total_ftnr += cint(row.ftnr)

        self.total_ftr = total_ftr
        self.total_ftnr = total_ftnr
        self.grand_total = total_ftr + total_ftnr

        if self.grand_total > 0:
            self.ftr_percentage = flt((total_ftr / self.grand_total) * 100, 2)
            self.ftnr_percentage = flt(
                (total_ftnr / self.grand_total) * 100, 2
            )
        else:
            self.ftr_percentage = 0.0
            self.ftnr_percentage = 0.0

        # Child table: table_zero_ip_funding Total Calculation
        total_zero_ip = 0
        for row in self.get("table_zero_ip_funding", []):
            total_zero_ip += cint(row.zero_ip_funding)

        self.zero_ip_funding_count = total_zero_ip


# ------------------------------------------------------------------
# TEMPLATE DOWNLOAD APIS (HEADER ONLY)
# ------------------------------------------------------------------
@frappe.whitelist()
def download_ftr_ftnr_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FTR_FTNR_Template"

    # Header Only
    ws.append(["Sol ID", "Date", "Status"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "Account_Opening_FTR_FTNR_Template.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"


@frappe.whitelist()
def download_zero_ip_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zero_IP_Template"

    # Header Only
    ws.append(["SOL ID", "A/C Opening Date", "Scheme Code"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "Zero_IP_Funding_Template.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "binary"


# ------------------------------------------------------------------
# 1ST EXCEL PROCESSOR (FTR / FTNR TRACKER)
# ------------------------------------------------------------------
@frappe.whitelist()
def process_consolidated_excel(file_url=None, confirm=False):
    if not file_url:
        frappe.throw("No file provided for processing.")

    confirm = frappe.parse_json(confirm)

    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_content = file_doc.get_content()

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        # Fetch Valid SOL IDs
        branches = frappe.get_all("Sahayog Branch", fields=["name"])
        valid_sol_ids = {
            str(b.get("name")).strip() for b in branches if b.get("name")
        }

        if frappe.db.has_column("Sahayog Branch", "sol_id"):
            extra_branches = frappe.get_all("Sahayog Branch", fields=["sol_id"])
            for b in extra_branches:
                if b.get("sol_id"):
                    valid_sol_ids.add(str(b.get("sol_id")).strip())

        # Header Validation
        headers = [
            str(cell.value).strip().lower() if cell.value is not None else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        if len(headers) < 3:
            frappe.throw(
                "<b>Excel Processing Failed:</b> File must have at least 3"
                " columns: <b>SOL ID, Date, Status</b>"
            )

        if not (
            "sol" in headers[0]
            and (
                "date" in headers[1]
                or "filled" in headers[1]
                or "form" in headers[1]
            )
            and "status" in headers[2]
        ):
            frappe.throw(
                f"<b>Excel Processing Failed - Invalid Column Headers!</b><br><br>"
                f"Expected Header Order:<br>"
                f"Column 1: <b>SOL ID</b> (Found: '{headers[0]}')<br>"
                f"Column 2: <b>Date / Form Filled</b> (Found: '{headers[1]}')<br>"
                f"Column 3: <b>Status</b> (Found: '{headers[2]}')"
            )

        errors = []
        valid_rows_data = []

        # Row by Row Validation
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            sol_id_raw = row[0] if len(row) > 0 else None
            date_raw = row[1] if len(row) > 1 else None
            status_raw = row[2] if len(row) > 2 else None

            if sol_id_raw is None and date_raw is None and status_raw is None:
                continue

            if sol_id_raw is None or date_raw is None or status_raw is None:
                errors.append(
                    f"Row {row_idx}: Missing value in SOL ID, Date, or Status."
                )
                continue

            sol_id = str(sol_id_raw).strip()
            if sol_id.endswith(".0"):
                sol_id = sol_id[:-2]

            status_clean = (
                str(status_raw)
                .strip()
                .upper()
                .replace(" ", "")
                .replace("\n", "")
                .replace("\r", "")
            )

            if sol_id not in valid_sol_ids:
                errors.append(
                    f"Row {row_idx}: SOL ID '<b>{sol_id}</b>' is not found in"
                    " <b>Sahayog Branch</b> DocType."
                )

            if status_clean not in ["FTR", "FTNR"]:
                errors.append(
                    f"Row {row_idx}: Invalid Status '<b>{status_raw}</b>'. Must"
                    " strictly be <b>FTR</b> or <b>FTNR</b>."
                )

            try:
                entry_date = getdate(date_raw)
                date_str = str(entry_date)
                month_name = entry_date.strftime("%B")
                year_num = entry_date.year
            except Exception:
                errors.append(
                    f"Row {row_idx}: Invalid Date format '<b>{date_raw}</b>'."
                )
                continue

            valid_rows_data.append({
                "sol_id": sol_id,
                "date_str": date_str,
                "month_name": month_name,
                "year_num": year_num,
                "status": status_clean,
            })

        if errors:
            error_list_html = "<br>".join(errors[:15])
            if len(errors) > 15:
                error_list_html += f"<br>...and {len(errors) - 15} more errors."

            frappe.throw(
                f"<b>File Upload Rejected! ({len(errors)} error(s)"
                " found)</b><br><br>"
                f"<div style='max-height: 200px;"
                f" overflow-y: auto;'>{error_list_html}</div>"
            )

        # Aggregate Data
        excel_grouped = {}
        for r in valid_rows_data:
            doc_key = (
                r["sol_id"],
                r["month_name"].capitalize(),
                str(r["year_num"]),
            )

            if doc_key not in excel_grouped:
                excel_grouped[doc_key] = {}

            if r["date_str"] not in excel_grouped[doc_key]:
                excel_grouped[doc_key][r["date_str"]] = {"ftr": 0, "ftnr": 0}

            if r["status"] == "FTR":
                excel_grouped[doc_key][r["date_str"]]["ftr"] += 1
            elif r["status"] == "FTNR":
                excel_grouped[doc_key][r["date_str"]]["ftnr"] += 1

        change_summary = []
        created_or_updated_docs = []

        for (sol_id, month, year), new_date_dict in excel_grouped.items():
            doc_name = f"{sol_id}-{month}-{year}"
            is_new = False

            if frappe.db.exists("Account Opening Operations", doc_name):
                doc = frappe.get_doc("Account Opening Operations", doc_name)
            else:
                doc = frappe.new_doc("Account Opening Operations")
                doc.sol_id = sol_id
                doc.month = month
                doc.year = year
                is_new = True

            existing_rows_map = {
                str(getdate(row_item.date)): row_item
                for row_item in doc.get("table_dllf", [])
                if row_item.date
            }
            doc_changes = []

            for date_val, counts in new_date_dict.items():
                if date_val in existing_rows_map:
                    row_obj = existing_rows_map[date_val]
                    old_ftr = cint(row_obj.ftr)
                    old_ftnr = cint(row_obj.ftnr)

                    if old_ftr != counts["ftr"] or old_ftnr != counts["ftnr"]:
                        doc_changes.append(
                            f"• Date <b>{date_val}</b>: FTR ({old_ftr} ➔"
                            f" <b>{counts['ftr']}</b>), FTNR ({old_ftnr} ➔"
                            f" <b>{counts['ftnr']}</b>)"
                        )
                        if confirm:
                            row_obj.ftr = counts["ftr"]
                            row_obj.ftnr = counts["ftnr"]
                else:
                    doc_changes.append(
                        f"• Date <b>{date_val}</b> [NEW ROW]: FTR ="
                        f" <b>{counts['ftr']}</b>, FTNR = <b>{counts['ftnr']}</b>"
                    )
                    if confirm:
                        doc.append(
                            "table_dllf",
                            {
                                "date": date_val,
                                "ftr": counts["ftr"],
                                "ftnr": counts["ftnr"],
                            },
                        )

            if is_new:
                change_summary.append(
                    f"<b>[NEW DOC] Record '{doc_name}' will be created</b> with"
                    f" {len(new_date_dict)} date entries."
                )
            elif doc_changes:
                change_summary.append(
                    f"<b>[UPDATE] Record '{doc_name}'</b> will be"
                    ' updated:<br>' + "<br>".join(doc_changes)
                )

            if confirm and (is_new or doc_changes):
                doc.save()
                created_or_updated_docs.append(doc.name)

        if confirm:
            frappe.db.commit()
            if not created_or_updated_docs:
                return {
                    "status": "no_change",
                    "message": (
                        "No updates made! All records were up to date."
                    ),
                }
            return {"status": "success", "docs": created_or_updated_docs}

        if not change_summary:
            return {
                "status": "no_change",
                "message": (
                    "No changes detected! All records in the Excel file are"
                    " already up to date."
                ),
            }

        return {
            "status": "requires_confirmation",
            "summary_html": "<br><br>".join(change_summary),
        }

    except frappe.ValidationError:
        raise
    except Exception as e:
        frappe.log_error(
            title="Excel Processing Error", message=frappe.get_traceback()
        )
        frappe.throw(
            f"An unexpected error occurred while processing Excel: {str(e)}"
        )


# ------------------------------------------------------------------
# 2ND EXCEL PROCESSOR (ZERO IP FUNDING TRACKER - table_zero_ip_funding)
# ------------------------------------------------------------------
@frappe.whitelist()
def process_zero_ip_excel(file_url=None, confirm=False):
    if not file_url:
        frappe.throw("No file provided for processing.")

    confirm = frappe.parse_json(confirm)

    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_content = file_doc.get_content()

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        # Fetch Valid SOL IDs
        branches = frappe.get_all("Sahayog Branch", fields=["name"])
        valid_sol_ids = {
            str(b.get("name")).strip() for b in branches if b.get("name")
        }

        if frappe.db.has_column("Sahayog Branch", "sol_id"):
            extra_branches = frappe.get_all("Sahayog Branch", fields=["sol_id"])
            for b in extra_branches:
                if b.get("sol_id"):
                    valid_sol_ids.add(str(b.get("sol_id")).strip())

        headers = [
            str(cell.value).strip().lower() if cell.value is not None else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        if len(headers) < 3:
            frappe.throw(
                "<b>Excel Processing Failed:</b> File must have at least 3"
                " columns: <b>SOL ID, A/C Opening Date, Scheme Code</b>"
            )

        errors = []
        valid_rows_data = []

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            sol_id_raw = row[0] if len(row) > 0 else None
            date_raw = row[1] if len(row) > 1 else None
            scheme_code_raw = row[2] if len(row) > 2 else None

            if (
                sol_id_raw is None
                and date_raw is None
                and scheme_code_raw is None
            ):
                continue

            if sol_id_raw is None or date_raw is None:
                errors.append(
                    f"Row {row_idx}: Missing value in SOL ID or A/C Opening"
                    " Date."
                )
                continue

            sol_id = str(sol_id_raw).strip()
            if sol_id.endswith(".0"):
                sol_id = sol_id[:-2]

            if sol_id not in valid_sol_ids:
                errors.append(
                    f"Row {row_idx}: SOL ID '<b>{sol_id}</b>' is not found in"
                    " <b>Sahayog Branch</b> DocType."
                )

            try:
                entry_date = getdate(date_raw)
                date_str = str(entry_date)
                month_name = entry_date.strftime("%B")
                year_num = entry_date.year
            except Exception:
                errors.append(
                    f"Row {row_idx}: Invalid Date format '<b>{date_raw}</b>'."
                )
                continue

            scheme_val = str(scheme_code_raw or "N/A").strip()

            valid_rows_data.append({
                "sol_id": sol_id,
                "ac_opening_date": date_str,
                "month_name": month_name,
                "year_num": year_num,
                "scheme_code": scheme_val if scheme_val else "N/A",
            })

        if errors:
            error_list_html = "<br>".join(errors[:15])
            if len(errors) > 15:
                error_list_html += f"<br>...and {len(errors) - 15} more errors."

            frappe.throw(
                "<b>Zero IP Funding File Upload Rejected! ("
                f"{len(errors)} error(s) found)</b><br><br>"
                f"<div style='max-height: 200px;"
                f" overflow-y: auto;'>{error_list_html}</div>"
            )

        # Structure: doc_key -> date -> scheme_code -> count
        excel_date_grouped = {}
        for r in valid_rows_data:
            doc_key = (
                r["sol_id"],
                r["month_name"].capitalize(),
                str(r["year_num"]),
            )
            date_key = r["ac_opening_date"]
            scheme_key = r["scheme_code"]

            if doc_key not in excel_date_grouped:
                excel_date_grouped[doc_key] = defaultdict(
                    lambda: defaultdict(int)
                )

            excel_date_grouped[doc_key][date_key][scheme_key] += 1

        change_summary = []
        created_or_updated_docs = []

        for (sol_id, month, year), date_dict in excel_date_grouped.items():
            doc_name = f"{sol_id}-{month}-{year}"
            is_new = False

            if frappe.db.exists("Account Opening Operations", doc_name):
                doc = frappe.get_doc("Account Opening Operations", doc_name)
            else:
                doc = frappe.new_doc("Account Opening Operations")
                doc.sol_id = sol_id
                doc.month = month
                doc.year = year
                is_new = True

            doc_changes = []

            # Map existing rows by Date
            existing_date_map = {
                str(getdate(row_item.ac_opening_date)): row_item
                for row_item in doc.get("table_zero_ip_funding", [])
                if row_item.ac_opening_date
            }

            for ac_date, schemes_map in date_dict.items():
                # Sort schemes descending by count
                sorted_schemes = sorted(
                    schemes_map.items(), key=lambda x: x[1], reverse=True
                )
                formatted_schemes = [
                    f"{sch}({cnt})" for sch, cnt in sorted_schemes
                ]
                top_scheme = (
                    formatted_schemes[0] if formatted_schemes else ""
                )
                options_json = json.dumps(formatted_schemes)
                total_date_count = sum(schemes_map.values())

                if ac_date in existing_date_map:
                    row_obj = existing_date_map[ac_date]
                    old_count = cint(row_obj.zero_ip_funding)

                    if (
                        old_count != total_date_count
                        or row_obj.scheme_code_options != options_json
                    ):
                        doc_changes.append(
                            f"• Date <b>{ac_date}</b>: Total Count ({old_count}"
                            f" ➔ <b>{total_date_count}</b>), Schemes:"
                            f" {', '.join(formatted_schemes)}"
                        )
                        if confirm:
                            # Preserve user selection if valid, else pick top scheme
                            if (
                                not row_obj.scheme_code
                                or row_obj.scheme_code not in formatted_schemes
                            ):
                                row_obj.scheme_code = top_scheme
                            row_obj.scheme_code_options = options_json
                            row_obj.zero_ip_funding = total_date_count
                else:
                    doc_changes.append(
                        f"• Date <b>{ac_date}</b> [NEW ROW]: Total Count ="
                        f" <b>{total_date_count}</b>, Schemes:"
                        f" {', '.join(formatted_schemes)}"
                    )
                    if confirm:
                        doc.append(
                            "table_zero_ip_funding",
                            {
                                "ac_opening_date": ac_date,
                                "scheme_code": top_scheme,
                                "scheme_code_options": options_json,
                                "zero_ip_funding": total_date_count,
                            },
                        )

            if is_new:
                change_summary.append(
                    f"<b>[NEW DOC] Record '{doc_name}' will be created</b> with"
                    f" {len(date_dict)} entries."
                )
            elif doc_changes:
                change_summary.append(
                    f"<b>[UPDATE] Record '{doc_name}'</b> will be"
                    ' updated:<br>' + "<br>".join(doc_changes)
                )

            if confirm and (is_new or doc_changes):
                doc.save()
                created_or_updated_docs.append(doc.name)

        if confirm:
            frappe.db.commit()
            if not created_or_updated_docs:
                return {
                    "status": "no_change",
                    "message": (
                        "No updates made! All records were up to date."
                    ),
                }
            return {"status": "success", "docs": created_or_updated_docs}

        if not change_summary:
            return {
                "status": "no_change",
                "message": (
                    "No changes detected! All records in the Excel file are"
                    " already up to date."
                ),
            }

        return {
            "status": "requires_confirmation",
            "summary_html": "<br><br>".join(change_summary),
        }

    except frappe.ValidationError:
        raise
    except Exception as e:
        frappe.log_error(
            title="Zero IP Excel Processing Error",
            message=frappe.get_traceback(),
        )
        frappe.throw(
            "An unexpected error occurred while processing Zero IP Excel:"
            f" {str(e)}"
        )