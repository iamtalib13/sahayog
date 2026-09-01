import frappe
import openpyxl
import io
import re
from datetime import datetime, date
from collections import Counter
from frappe.model.document import Document
from frappe.utils import cint, getdate
from frappe.utils.response import build_response


class Miscellaneous(Document):

    def autoname(self):
        # Format: 1005-August-2026
        if self.sol_id and self.month and self.year:
            self.name = f"{self.sol_id}-{self.month.capitalize()}-{self.year}"

    def validate(self):
        self.sort_child_tables_by_date()
        self.calculate_totals()

    def sort_child_tables_by_date(self):
        # Helper date resolution logic for child rows
        def resolve_row_date(row):
            d_val = getattr(row, "account_opening_error_date", None) or getattr(row, "error_date", None)
            return getdate(str(d_val)) if d_val else getdate("1900-01-01")

        # Sort Account Opening Error Child Table Descending
        if self.get("account_opening_error"):
            sorted_rows = sorted(
                self.get("account_opening_error"),
                key=resolve_row_date,
                reverse=True,
            )
            self.set("account_opening_error", [])
            for idx, row in enumerate(sorted_rows, start=1):
                row.idx = idx
                self.append("account_opening_error", row)

        # Sort Bank Reconciliation Discrepancy Child Table Descending
        if self.get("bank_reconciliation_discrepancy"):
            sorted_rows = sorted(
                self.get("bank_reconciliation_discrepancy"),
                key=resolve_row_date,
                reverse=True,
            )
            self.set("bank_reconciliation_discrepancy", [])
            for idx, row in enumerate(sorted_rows, start=1):
                row.idx = idx
                self.append("bank_reconciliation_discrepancy", row)

    def calculate_totals(self):
        total_account_errors = sum(
            cint(row.error_count) for row in self.get("account_opening_error", [])
        )
        self.account_opening_error_count = total_account_errors

        total_reco_discrepancies = sum(
            cint(row.error_count)
            for row in self.get("bank_reconciliation_discrepancy", [])
        )
        self.reconciliation_discrepancy_count = total_reco_discrepancies


# ------------------------------------------------------------------
# MULTI-DATE PARSER HELPER
# ------------------------------------------------------------------
def extract_dates_from_cell(cell_value):
    """
    Parses single cells containing multiple dates (separated by ',', '&', 'and', '|', newline)
    and returns a list of parsed datetime.date objects.
    """
    if cell_value is None:
        return []

    # If cell value is already datetime/date instance
    if isinstance(cell_value, (datetime, date)):
        return [getdate(cell_value)]

    raw_str = str(cell_value).strip()
    if not raw_str:
        return []

    # Split string by common delimiters: comma, &, 'and', vertical bar, newlines
    tokens = re.split(r'\s*(?:&|,|\band\b|\||\n)\s*', raw_str, flags=re.IGNORECASE)
    parsed_dates = []

    for token in tokens:
        clean_token = token.strip()
        if not clean_token:
            continue
        try:
            d_obj = getdate(clean_token)
            if d_obj:
                parsed_dates.append(d_obj)
        except Exception:
            # Fallback format checking if standard getdate fails
            for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d.%m.%Y'):
                try:
                    d_obj = datetime.strptime(clean_token, fmt).date()
                    parsed_dates.append(d_obj)
                    break
                except ValueError:
                    pass

    return parsed_dates


# ------------------------------------------------------------------
# EXCEL PROCESSOR FOR MISCELLANEOUS DOCTYPE
# ------------------------------------------------------------------
@frappe.whitelist()
def process_miscellaneous_excel(file_url=None, type_name=None, confirm=False):
    if not file_url:
        frappe.throw("No file provided for processing.")

    if not type_name:
        frappe.throw("Please specify the Type Name (e.g., 'Account Opening Error').")

    confirm = frappe.parse_json(confirm)

    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_content = file_doc.get_content()

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        sheet = wb.active

        # Fetch Valid SOL IDs from Sahayog Branch
        branches = frappe.get_all("Sahayog Branch", fields=["name"])
        valid_sol_ids = {
            str(b.get("name")).strip() for b in branches if b.get("name")
        }

        if frappe.db.has_column("Sahayog Branch", "sol_id"):
            extra_branches = frappe.get_all("Sahayog Branch", fields=["sol_id"])
            for b in extra_branches:
                if b.get("sol_id"):
                    valid_sol_ids.add(str(b.get("sol_id")).strip())

        # ------------------------------------------------------------------
        # STRICT HEADER VALIDATION (EXACT HEADERS CHECK)
        # ------------------------------------------------------------------
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(cell).strip() for cell in first_row if cell is not None and str(cell).strip() != ""]

        if type_name == "Account Opening Error":
            expected_c1 = "BRANCH CODE ( SOL ID )"
            expected_c2 = "A/C Opening Date"
        else:
            expected_c1 = "Sol Id"
            expected_c2 = "Error Date"

        if len(headers) != 2:
            frappe.throw(
                f"<b>Excel Format Rejected!</b><br><br>"
                f"Your file contains <b>{len(headers)} columns</b>.<br>"
                f"Required Format: <b>EXACTLY 2 COLUMNS</b>.<br><br>"
                f"Required Column Structure:<br>"
                f"• Column 1: <b>{expected_c1}</b><br>"
                f"• Column 2: <b>{expected_c2}</b>"
            )

        col1_name = headers[0].strip()
        col2_name = headers[1].strip()

        if col1_name.lower() != expected_c1.lower() or col2_name.lower() != expected_c2.lower():
            frappe.throw(
                f"<b>Invalid Column Headers!</b><br><br>"
                f"Excel headers must be strictly <b>'{expected_c1}'</b> and <b>'{expected_c2}'</b>.<br><br>"
                f"Found Headers:<br>"
                f"• Column 1: '<b>{headers[0]}</b>' (Expected: <b>{expected_c1}</b>)<br>"
                f"• Column 2: '<b>{headers[1]}</b>' (Expected: <b>{expected_c2}</b>)"
            )

        errors = []
        valid_rows_data = []

        # Row-by-Row Validation
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            sol_id_raw = row[0] if len(row) > 0 else None
            date_raw = row[1] if len(row) > 1 else None

            if sol_id_raw is None and date_raw is None:
                continue

            if sol_id_raw is None or date_raw is None:
                errors.append(
                    f"Row {row_idx}: Missing value in {expected_c1} or {expected_c2} column."
                )
                continue

            sol_id = str(sol_id_raw).strip()
            if sol_id.endswith(".0"):
                sol_id = sol_id[:-2]

            if sol_id not in valid_sol_ids:
                errors.append(
                    f"Row {row_idx}: SOL ID '<b>{sol_id}</b>' is not found in"
                    " <b>Sahayog Branch</b> DocType."
                )

            parsed_dates = extract_dates_from_cell(date_raw)
            if not parsed_dates:
                errors.append(
                    f"Row {row_idx}: Invalid or unparseable Date format '<b>{date_raw}</b>'."
                )
                continue

            for entry_date in parsed_dates:
                date_str = str(entry_date)
                month_name = entry_date.strftime("%B")
                year_num = entry_date.year

                valid_rows_data.append({
                    "sol_id": sol_id,
                    "error_date": date_str,
                    "month_name": month_name,
                    "year_num": year_num,
                })

        if errors:
            error_list_html = "<br>".join(errors[:15])
            if len(errors) > 15:
                error_list_html += f"<br>...and {len(errors) - 15} more errors."

            frappe.throw(
                f"<b>File Upload Rejected! ({len(errors)} error(s)"
                " found)</b><br><br>"
                f"<div style='max-height: 200px;"
                f" overflow-y: auto;'>{error_list_html}</div>"
            )

        # Structure: doc_key (sol_id, month, year) -> date -> count
        excel_grouped = {}
        for r in valid_rows_data:
            doc_key = (
                r["sol_id"],
                r["month_name"].capitalize(),
                str(r["year_num"]),
            )

            if doc_key not in excel_grouped:
                excel_grouped[doc_key] = Counter()

            excel_grouped[doc_key][r["error_date"]] += 1

        # Target Child Table mapping
        if type_name == "Account Opening Error":
            child_table_field = "account_opening_error"
        elif type_name == "Bank Reconciliation Discrepancy":
            child_table_field = "bank_reconciliation_discrepancy"
        else:
            frappe.throw(f"Unsupported Type Name: {type_name}")

        # FIX 1: Scheme json ke mutabiq child table date field name 'error_date' hai
        date_field_name = "error_date"

        change_summary = []
        created_or_updated_docs = []

        for (sol_id, month, year), date_counts in excel_grouped.items():
            doc_name = f"{sol_id}-{month}-{year}"
            is_new = False

            if frappe.db.exists("Miscellaneous", doc_name):
                doc = frappe.get_doc("Miscellaneous", doc_name)
            else:
                doc = frappe.new_doc("Miscellaneous")
                doc.sol_id = sol_id
                doc.month = month
                doc.year = cint(year) # FIX 2: Integer type conversion
                doc.name = doc_name   # FIX 3: Explicit name set
                is_new = True

            # Existing rows mapping
            existing_rows_map = {}
            for row_item in doc.get(child_table_field, []):
                item_date = getattr(row_item, date_field_name, None)
                if item_date:
                    existing_rows_map[str(getdate(item_date))] = row_item

            doc_changes = []

            for date_val, new_count in date_counts.items():
                if date_val in existing_rows_map:
                    row_obj = existing_rows_map[date_val]
                    old_count = cint(row_obj.error_count)

                    if old_count != new_count:
                        doc_changes.append(
                            f"• Date <b>{date_val}</b>: Count ({old_count} ➔"
                            f" <b>{new_count}</b>)"
                        )
                        if confirm:
                            row_obj.error_count = new_count
                else:
                    doc_changes.append(
                        f"• Date <b>{date_val}</b> [NEW ROW]: Count = <b>{new_count}</b>"
                    )
                    if confirm:
                        doc.append(
                            child_table_field,
                            {
                                date_field_name: date_val,
                                "error_count": new_count,
                            },
                        )

            if is_new:
                change_summary.append(
                    f"<b>[NEW DOC] Record '{doc_name}' will be created</b> with"
                    f" {len(date_counts)} date entries."
                )
            elif doc_changes:
                change_summary.append(
                    f"<b>[UPDATE] Record '{doc_name}'</b> will be"
                    ' updated:<br>' + "<br>".join(doc_changes)
                )

            if confirm and (is_new or doc_changes):
                doc.save()
                created_or_updated_docs.append(doc.name)

        if confirm:
            frappe.db.commit()
            if not created_or_updated_docs:
                return {
                    "status": "no_change",
                    "message": "No updates made! All records were up to date.",
                }
            return {"status": "success", "docs": created_or_updated_docs}

        if not change_summary:
            return {
                "status": "no_change",
                "message": (
                    "No changes detected! All records in the Excel file are"
                    " already up to date."
                ),
            }

        return {
            "status": "requires_confirmation",
            "summary_html": "<br><br>".join(change_summary),
        }

    except frappe.ValidationError:
        raise
    except Exception as e:
        frappe.log_error(
            title="Miscellaneous Excel Processing Error", message=frappe.get_traceback()
        )
        frappe.throw(
            f"An unexpected error occurred while processing Excel: {str(e)}"
        )


# ------------------------------------------------------------------
# TEMPLATE GENERATOR FOR MISCELLANEOUS EXCEL
# ------------------------------------------------------------------
@frappe.whitelist()
def download_miscellaneous_template(type_name=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # Dynamic Column Setup
    if type_name == "Account Opening Error":
        headers = ["BRANCH CODE ( SOL ID )", "A/C Opening Date"]
    else:
        headers = ["Sol Id", "Error Date"]

    # Sirf Header add hoga (Sample rows remove kar diye hain)
    ws.append(headers)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 25

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"{type_name.replace(' ', '_')}_Template.xlsx" if type_name else "Miscellaneous_Template.xlsx"

    frappe.response['filename'] = filename
    frappe.response['filecontent'] = stream.getvalue()
    frappe.response['type'] = 'binary'
    return build_response()