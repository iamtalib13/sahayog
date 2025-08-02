# Copyright (c) 2025, Developer Team and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import _
import pandas as pd
import io
import re
import csv

class BDOPerformance(Document):
	pass

# Try to import openpyxl for Excel support
try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

@frappe.whitelist()
def get_user_mis_data():
    """
    Get logged-in user's data from MIS Report attachment using Employee ID
    Supports both CSV and Excel files
    """
    try:
        current_user = frappe.session.user
        
        # Special handling for Administrator
        if current_user == 'Administrator':
            return 
        
        # Extract Employee ID from user email
        emp_id = extract_emp_id_from_email(current_user)
        
        if not emp_id:
            return {
                'success': False,
                'message': f'Could not extract Employee ID from user: {current_user}'
            }
        
        # Get the MIS Report record
        mis_report = frappe.get_all(
            'MIS Report',
            fields=['name', 'report_name', 'report_attachment', 'last_updated_date'],
            filters={
                'analytics_chart': 'BDO Performance',
                'is_active': 1
            },
            order_by='last_updated_date desc',
            limit=1
        )
        
        if not mis_report or not mis_report[0].report_attachment:
            return {'success': False, 'message': 'No BDO Performance report or attachment found'}
        
        report = mis_report[0]
        
        # Get file content
        file_doc = frappe.get_doc('File', {'file_url': report.report_attachment})
        file_content = file_doc.get_content()
        
        # Handle different file types
        if report.report_attachment.lower().endswith('.csv'):
            return parse_csv_for_user(file_content, emp_id, report, current_user)
        elif report.report_attachment.lower().endswith(('.xlsx', '.xls')):
            if not EXCEL_SUPPORT:
                return {
                    'success': False,
                    'message': 'Excel support not available. Please install openpyxl or convert to CSV format.'
                }
            return parse_excel_for_user(file_content, emp_id, report, current_user)
        else:
            return {
                'success': False,
                'message': 'Unsupported file format. Please use CSV or Excel files.'
            }
        
    except Exception as e:
        frappe.log_error(f"Error in get_user_mis_data: {str(e)}")
        return {'success': False, 'message': f'Error: {str(e)}'}

def parse_csv_for_user(file_content, emp_id, report, current_user):
    """Parse CSV file and find user data"""
    try:
        content_str = file_content.decode('utf-8')
        lines = content_str.strip().split('\n')
        
        if len(lines) < 2:
            return {'success': False, 'message': 'CSV file is empty or has no data rows'}
        
        headers = [h.strip() for h in lines[0].split(',')]
        
        # Find employee ID column
        emp_id_col_index = find_emp_id_column(headers)
        if emp_id_col_index is None:
            return {
                'success': False,
                'message': 'No employee ID column found',
                'available_columns': headers,
                'extracted_emp_id': emp_id
            }
        
        # Find user data
        for line in lines[1:]:
            values = [v.strip() for v in line.split(',')]
            if len(values) > emp_id_col_index and str(values[emp_id_col_index]).strip() == str(emp_id):
                user_data = {}
                for i, header in enumerate(headers):
                    if i < len(values):
                        user_data[header] = values[i]
                
                return {
                    'success': True,
                    'data': user_data,
                    'user': current_user,
                    'emp_id': emp_id,
                    'report_name': report.report_name,
                    'last_updated': report.last_updated_date
                }
        
        return {
            'success': False,
            'message': f'No data found for Employee ID: {emp_id}',
            'available_columns': headers,
            'extracted_emp_id': emp_id
        }
        
    except Exception as e:
        return {'success': False, 'message': f'Error parsing CSV: {str(e)}'}

def parse_excel_for_user(file_content, emp_id, report, current_user):
    """Parse Excel file and find user data"""
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value if cell.value else '')
        
        # Find employee ID column
        emp_id_col_index = find_emp_id_column(headers)
        if emp_id_col_index is None:
            return {
                'success': False,
                'message': 'No employee ID column found',
                'available_columns': headers,
                'extracted_emp_id': emp_id
            }
        
        # Find user data
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) > emp_id_col_index and str(row[emp_id_col_index]).strip() == str(emp_id):
                user_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        user_data[header] = row[i] if row[i] is not None else ''
                
                return {
                    'success': True,
                    'data': user_data,
                    'user': current_user,
                    'emp_id': emp_id,
                    'report_name': report.report_name,
                    'last_updated': report.last_updated_date
                }
        
        return {
            'success': False,
            'message': f'No data found for Employee ID: {emp_id}',
            'available_columns': headers,
            'extracted_emp_id': emp_id
        }
        
    except Exception as e:
        return {'success': False, 'message': f'Error parsing Excel: {str(e)}'}

def find_emp_id_column(headers):
    """Find employee ID column index"""
    possible_emp_columns = ['emp_id', 'empid', 'employee_id', 'employee', 'id']
    for i, header in enumerate(headers):
        if header and header.lower() in possible_emp_columns:
            return i
    return None

def extract_emp_id_from_email(email):
    """Extract Employee ID from email address"""
    try:
        if '@' in email:
            username = email.split('@')[0]
            if username.isdigit():
                return username
            numbers = re.findall(r'\d+', username)
            if numbers:
                return numbers[0]
        return None
    except Exception as e:
        frappe.log_error(f"Error extracting emp_id from email {email}: {str(e)}")
        return None

# @frappe.whitelist()
# def get_employee_kpi():
#     user_email = frappe.session.user  # Or dynamically from user
#     if user_email == "Guest":
#         frappe.throw(_("You must be logged in to view this dashboard."))

#     # Extract Emp ID from email
#     emp_id = user_email.split("@")[0]  # "8446"


#     # Step 1: Check if the employee record exists in BDO Performance
#     result = frappe.db.sql("""
#         SELECT name FROM `tabBDO Performance`
#         WHERE empid = %s
#     """, (emp_id,), as_dict=True)

#     if not result:
#         frappe.throw(_("No BDO Performance record found for this employee."))

#     # Step 2: Fetch KPI data
#     data = frappe.db.sql("""
#         SELECT
#          zone_name,
# 			region_name,
# 			district_name,
# 			branch_name,
# 			empid,
# 			name1,
# 			designation_name,
# 			join_dur,
# 			inactive,
# 			active,
# 			total_ssagnt,
# 			new_ssagnt,
# 			new_rd_ac,
# 			new_rd_amt,
# 			new_smbg_ac,
# 			new_smbg_amt,
# 			total_rdsmbg_ac,
# 			total_rdsmbg_amt,
# 			dam_ac,
# 			dam_amt,
# 			fd_ac,
# 			fd_atm,
# 			fd_6m_ac,
# 			fd_6m_amt,
# 			mis_ac,
# 			mis_amt,
# 			total_fd_ac,
# 			total_fd_amt,
# 			rddemand,
# 			rdcolle,
# 			smbg_demand,
# 			smbg_colle,
# 			total_rdsmbg_demand,
#             modified,
# 			total_rdsmbg_collection                
#         FROM `tabBDO Performance`
#         WHERE empid = %s
#         LIMIT 1
#     """, (emp_id,), as_dict=True)[0]

#     # Step 3: Calculate % Achieved
#     demand = data.get("total_rdsmbg_demand") or 0
#     collected = data.get("total_rdsmbg_collection") or 0
#     data["percent_achieved"] = round((collected / demand) * 100) if demand else 0

#     return data
