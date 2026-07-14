import csv
import io
from datetime import datetime


import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt, cint, nowdate, getdate, get_first_day, get_last_day, date_diff
from frappe.utils import now_datetime


import frappe
import openpyxl
from frappe import _
from frappe.model.document import Document
from frappe.utils import (
    flt,
    fmt_money,
    format_datetime,
    get_first_day,
    get_last_day,
    getdate,
    now,
    nowdate
)
from frappe.utils.xlsxutils import make_xlsx
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from frappe.utils import getdate, nowdate, date_diff

# Custom App Imports
from sahayog.petty_cash_management.api.finacle_integration import individual_finacle_fund_transfer_api
from sahayog.petty_cash_management.permissions import get_user_allowed_branches


class PettyCashTransaction(Document):

    def before_insert(self):
        if not self.transaction_date:
            self.transaction_date = nowdate()

        # Default Status
        self.approval_status = "Draft"

        if "HO Petty Cash Manager" not in frappe.get_roles():
            self.transaction_type = "Expense"
            emp_branch = frappe.db.get_value(
                "Employee", {"user_id": frappe.session.user, "status": "Active"}, "sahayog_branch")
            if emp_branch:
                self.branch = emp_branch

    def before_save(self):
        if self.branch:
            self.branch_name = frappe.db.get_value(
                "Sahayog Branch", self.branch, "branch")

        if self.transaction_type == "Expense":
            total_expense = sum(flt(item.amount) for item in self.items)
            self.amount = total_expense

            # Recalculate Limits Breakdown on every save
            self.calculate_limit_breakdown()

        if self.transaction_type == "Expense":
            total_expense = sum(flt(item.amount) for item in self.items)
            self.amount = total_expense

            # [NEW] Generate GL Codes dynamically
            self.generate_item_gl_codes()

            self.calculate_limit_breakdown()

        # [NEW ADDITION] Auto-set Source Account for Fund Allocation
        if self.transaction_type == "Fund Allocation":
            self.set_default_source_account()

            # [NEW] Sync Updates to Journal Entry
            # Only if Draft (docstatus 0) and JE already exists
            if self.docstatus == 0 and self.journal_entry_ref:
                self.sync_journal_entry_changes()

        # [NEW] Prevent modification if attempt was made
        if self.submission_attempted:
            # Check if critical fields changed using db_doc comparison
            db_doc = frappe.db.get_value("Petty Cash Transaction", self.name,
                                         ["amount", "target_scope"], as_dict=True)

            if db_doc and (flt(self.amount) != flt(db_doc.amount) or self.target_scope != db_doc.target_scope):
                frappe.throw(
                    _("Cannot modify Amount or Scope after a submission attempt has been made."))

    def sync_journal_entry_changes(self):
        """
        [NEW] Updates the existing Journal Entry instead of deleting it.
        """
        # 1. Check if critical fields changed
        db_doc = frappe.db.get_value("Petty Cash Transaction", self.name,
                                     ["amount", "is_bulk_allocation", "target_scope", "source_bank_account", "branch"], as_dict=True)

        if not db_doc:
            return

        has_changed = (
            flt(self.amount) != flt(db_doc.amount) or
            self.is_bulk_allocation != db_doc.is_bulk_allocation or
            self.target_scope != db_doc.target_scope or
            self.source_bank_account != db_doc.source_bank_account or
            self.branch != db_doc.branch
        )

        if has_changed and self.journal_entry_ref:
            # Switch to Admin to edit JE
            original_user = frappe.session.user
            frappe.set_user("Administrator")

            try:
                # 2. Load Existing JE
                je = frappe.get_doc("Journal Entry", self.journal_entry_ref)

                # 3. Clear existing Accounts table
                je.accounts = []

                # 4. Re-calculate Accounts (Same logic as create_ho_fund_allocation_je)
                # --- LOGIC REUSE START ---
                target_wallets = []
                if self.is_bulk_allocation:
                    filters = {"status": "Active", "is_fund_source": 0}
                    if self.target_scope == "Metro Branches Only":
                        filters["branch_type"] = "Metro"
                    elif self.target_scope == "Non-Metro Branches Only":
                        filters["branch_type"] = "Non Metro"
                    target_wallets = frappe.get_all("Branch Petty Cash Account", filters=filters, fields=[
                                                    "name", "branch", "gl_sub_code"])
                else:
                    target_wallets = frappe.get_all("Branch Petty Cash Account", filters={
                                                    "branch": self.branch}, fields=["name", "branch", "gl_sub_code"])

                total_amount = 0.0
                cost_center = frappe.get_cached_value(
                    'Company', je.company, 'cost_center')

                for wallet in target_wallets:
                    if not wallet.gl_sub_code:
                        continue
                    acc_name = frappe.db.sql(
                        "SELECT name FROM `tabAccount` WHERE account_number=%s", wallet.gl_sub_code)
                    if not acc_name:
                        continue
                    acc_name = acc_name[0][0]

                    je.append("accounts", {
                        "account": acc_name,
                        "debit_in_account_currency": 0,
                        "credit_in_account_currency": self.amount,
                        "cost_center": cost_center,
                        "user_remark": f"Allocation to {wallet.branch}"
                    })
                    total_amount += self.amount

                # Debit Line
                je.append("accounts", {
                    "account": self.source_bank_account,
                    "debit_in_account_currency": total_amount,
                    "credit_in_account_currency": 0,
                    "cost_center": cost_center,
                    "user_remark": f"Total Fund Allocation for {len(target_wallets)} branches"
                })
                # --- LOGIC REUSE END ---

                # 5. Save the JE
                je.flags.ignore_permissions = True
                je.save(ignore_permissions=True)

                frappe.msgprint(
                    _("Linked Journal Entry updated successfully."))

            except Exception as e:
                frappe.log_error(title="JE Sync Error",
                                 message=frappe.get_traceback())
                frappe.msgprint(
                    _("Could not sync Journal Entry changes. Please check error log."))

            finally:
                frappe.set_user(original_user)

    def generate_item_gl_codes(self):
        if not self.branch or not self.items:
            return

        # 1. Collect all Category IDs from the child table rows
        # 'item.expense_category' stores the ID (e.g., '1004'), not the name
        category_ids = [
            item.expense_category for item in self.items if item.expense_category]

        # 2. Fetch Suffixes for these IDs
        # Returns dict: {'1004': '01840390001', '1005': '...'}
        category_map = {}
        if category_ids:
            # We filter by 'name' (the ID) because that's what is stored in the link field
            results = frappe.get_all(
                "Expense Category",
                filters={"name": ["in", category_ids]},
                fields=["name", "finacle_gl_code"],
                as_list=True
            )
            category_map = {row[0]: row[1] for row in results}

        # 3. Apply Logic: Branch Code + Suffix
        for item in self.items:
            # Get suffix using the Link field value (ID)
            suffix = category_map.get(item.expense_category)

            if suffix:
                item.finacle_gl_code = f"{self.branch}{suffix}"
            else:
                item.finacle_gl_code = ""

    def set_default_source_account(self):
        """
        [NEW] Automatically finds the HO Branch (is_fund_source=1) and sets its GL Account 
        into the 'source_bank_account' field.
        """
        if self.source_bank_account:
            return  # Already set manually, skipping auto-set

        # Find the Wallet marked as Source
        ho_wallet = frappe.db.get_value("Branch Petty Cash Account",
                                        {"is_fund_source": 1, "status": "Active"},
                                        ["name", "gl_sub_code"],
                                        as_dict=True
                                        )

        if not ho_wallet:
            # Optional: You can choose to throw an error or just let validation fail later
            # frappe.throw(_("No Active Branch Account found with 'Is Fund Source' checked."))
            return

        # Find the Chart of Accounts Name for this GL Code
        # We use SQL to bypass permission checks
        account_name = frappe.db.sql(
            """SELECT name FROM `tabAccount` WHERE account_number=%s""", ho_wallet.gl_sub_code)

        if account_name:
            self.source_bank_account = account_name[0][0]

    def validate(self):

        # [NEW] Validate Item Descriptions (Max 30 chars)
        self.validate_item_descriptions()

        # REQUIREMENT 1: On Save, create Draft Journal Entry if Fund Allocation
        # if self.transaction_type == "Fund Allocation" and not self.journal_entry_ref:
        #     self.create_ho_fund_allocation_je()

        if self.transaction_type == "Fund Allocation":
            if not self.amount or self.amount <= 0:
                frappe.throw(_("Amount is required for Fund Allocation"))

            # ADD THIS BLOCK:
    # If this is coming from the Finacle Sync, DO NOT create the Journal Entry
            # if self.posted_to_finacle:
            #     return

            # if not self.journal_entry_ref:
            #     self.create_ho_fund_allocation_je()

           # [FIXED LOGIC HERE]
            # We check two things:
            # 1. Is it NOT posted to finacle? (If posted=1, we skip this)
            # 2. Is journal_entry_ref empty? (If yes, we need to create one)
            if not self.posted_to_finacle and not self.journal_entry_ref:
                self.create_ho_fund_allocation_je()

        # 1. Check Account Existence
        if not frappe.db.exists("Branch Petty Cash Account", {"branch": self.branch}):
            frappe.throw(
                _("Branch Petty Cash Account for branch '{0}' not found!").format(self.branch))

        wallet = frappe.get_doc("Branch Petty Cash Account", {
                                "branch": self.branch})

        if self.transaction_type == "Expense":
            if not self.items:
                frappe.throw(_("At least one expense item is required."))

            self.amount = sum(flt(item.amount) for item in self.items)
            self.calculate_limit_breakdown()
            self.validate_bill_dates()
            # self.validate_expense_soft(wallet)
            self.validate_expense_against_active_wallet()

        # [FIX] Fetch BOTH Real Balance and Unsettled Cash for display
        wallet_values = frappe.db.get_value("Branch Petty Cash Account",
                                            {"branch": self.branch},
                                            ["current_balance", "unsettled_cash"],
                                            as_dict=True
                                            )

        if wallet_values:
            self.current_branch_balance = flt(wallet_values.current_balance)
            self.current_unsettled_cash = flt(
                wallet_values.unsettled_cash)  # <--- New Field

    def validate_bill_dates(self):
        today = getdate(nowdate())

        for row in self.items:
            if row.bill_date:
                bill_date = getdate(row.bill_date)
                diff = date_diff(today, bill_date)

                if diff < 0:
                    frappe.throw(
                        f"Row #{row.idx}: Bill Date cannot be in the future."
                    )

                if diff > 30:
                    frappe.throw(
                        f"Row #{row.idx}: Bill Date cannot be older than 30 days from today."
                    )

    def calculate_limit_breakdown(self):
        """
        Calculates Limit.
        UPDATED: Verified transactions are excluded from 'Used Amount', 
        automatically replenishing the limit for the category.
        """
        if self.transaction_type != "Expense":
            return

        branch_type = frappe.db.get_value("Branch Petty Cash Account", {
                                          "branch": self.branch}, "branch_type")

        # 1. Summarize current doc's expenses
        current_tx_categories = {}
        for item in self.items:
            current_tx_categories.setdefault(item.expense_category, 0.0)
            current_tx_categories[item.expense_category] += flt(item.amount)

        total_within = 0.0
        total_exceeding = 0.0

        first_day = get_first_day(self.transaction_date)
        last_day = get_last_day(self.transaction_date)

        for category_id, tx_amount in current_tx_categories.items():
            category_doc = frappe.get_doc("Expense Category", category_id)
            limit = category_doc.metro_limit if branch_type == "Metro" else category_doc.non_metro_limit

            # Unlimited check
            if limit == 0:
                total_within += tx_amount
                continue

            # 2. Get Spent Amount
            # CRITICAL CHANGE: Added "AND parent.approval_status != 'Verified'"
            # This means verified bills stop counting as "Pending Usage".
            spent_sql = """
                SELECT COALESCE(SUM(child.amount), 0)
                FROM `tabPetty Cash Transaction Item` child
                JOIN `tabPetty Cash Transaction` parent ON child.parent = parent.name
                WHERE parent.branch = %s 
                  AND child.expense_category = %s
                  AND parent.transaction_date BETWEEN %s AND %s
                  AND parent.docstatus = 1
                  AND parent.approval_status != 'Verified'
                  AND parent.name != %s
            """

            already_spent = frappe.db.sql(
                spent_sql, (self.branch, category_id, first_day, last_day, self.name or "New"))[0][0]

            # 3. Calculate Remaining
            # Logic: Limit (3000) - Pending (0) = 3000.
            # It naturally stops at the limit (3000) and won't go to 4000.
            remaining_limit = max(flt(limit) - flt(already_spent), 0)

            if tx_amount <= remaining_limit:
                total_within += tx_amount
            else:
                can_spend = remaining_limit
                excess = tx_amount - remaining_limit
                total_within += can_spend
                total_exceeding += excess

        self.amount_within_limit = total_within
        self.amount_exceeding_limit = total_exceeding

    def validate_expense_soft(self, wallet):
        # 1. Fetch Latest Values Directly from DB (Bypassing any old calculation logic)
        wallet_data = frappe.db.get_value("Branch Petty Cash Account",
                                          {"branch": self.branch},
                                          ["current_balance", "unsettled_cash"],
                                          as_dict=True
                                          )

        if not wallet_data:
            return

        # 2. Read the Real Synced Balance
        available_balance = flt(wallet_data.current_balance)  # Should be 21047
        unsettled_cash = flt(wallet_data.unsettled_cash)     # Should be 12000

        # 3. Calculate Total Buying Power (Bank + Cash Hand)
        total_buying_power = available_balance + unsettled_cash

        # 4. Validate
        if total_buying_power < self.amount:
            frappe.throw(_("Insufficient Funds. Bank: ₹{0} + Cash-in-Hand: ₹{1} = Total: ₹{2}. Required: ₹{3}").format(
                available_balance, unsettled_cash, total_buying_power, self.amount))

        # 5. Check Category Limits
        if self.amount_exceeding_limit > 0:
            frappe.msgprint(_("Warning: Expenses exceed category limits by ₹{0}. This amount will NOT be deducted until approved by HO.").format(
                self.amount_exceeding_limit), alert=True)

    def on_submit(self):
        if self.transaction_type == "Expense":
            self.validate_expense_against_active_wallet()
        # 1. Fund Allocation flow
        if self.transaction_type == "Fund Allocation":
            if not self.postedtofinacle:
                self.processfinacletransfer()
            else:
                frappe.msgprint(
                    _("Fund Allocation Synced from Finacle successfully."))

            self.db_set("amount_deducted", 0, update_modified=False)
            self.db_set("approval_status", "Posted", update_modified=False)
            self.update_wallet()
            return

        # 2. Expense flow
        if self.transaction_type == "Expense":
            if self.amount_exceeding_limit > 0:
                self.db_set("amount_deducted",
                            self.amount_within_limit, update_modified=False)
                self.db_set("approval_status", "Pending Approval",
                            update_modified=False)
            else:
                self.db_set("amount_deducted", self.amount,
                            update_modified=False)
                self.db_set("approval_status", "Approved",
                            update_modified=False)

            if self.is_legacy_unsettled_cash_flow_enabled():
                wallet = self.get_branch_wallet_doc()
                wallet.update_unsettled_cash(self.amount, "Expense")
            else:
                self.apply_new_flow_wallet_deduction()

            enable_integration = cint(
                frappe.db.get_single_value(
                    "Sahayog Settings", "enable_finacle_integration") or 0
            )

            if enable_integration:
                self.createjournalentry()

            self.update_wallet()

    def create_ho_fund_allocation_je(self):
        """
        Creates Draft Journal Entry for Single or Bulk Allocation.
        """
        original_user = frappe.session.user
        frappe.set_user("Administrator")

        try:
            # 1. Identify Target Branches (Logic remains same)
            target_wallets = []
            if self.is_bulk_allocation:
                filters = {"status": "Active", "is_fund_source": 0}
                if self.target_scope == "Metro Branches Only":
                    filters["branch_type"] = "Metro"
                elif self.target_scope == "Non-Metro Branches Only":
                    filters["branch_type"] = "Non Metro"
                target_wallets = frappe.get_all("Branch Petty Cash Account", filters=filters, fields=[
                                                "name", "branch", "gl_sub_code"])
                if not target_wallets:
                    frappe.throw(
                        _("No branches found matching the selected Scope."))
            else:
                if not self.branch:
                    frappe.throw(
                        _("Branch is required for Single Allocation."))
                target_wallets = frappe.get_all("Branch Petty Cash Account", filters={
                                                "branch": self.branch}, fields=["name", "branch", "gl_sub_code"])

            # 2. Prepare JE Accounts List
            je_accounts = []
            total_amount = 0.0
            company = frappe.defaults.get_user_default(
                "Company") or frappe.db.get_single_value("Global Defaults", "default_company")
            cost_center = frappe.get_cached_value(
                'Company', company, 'cost_center')

            # Loop through targets (Credits)
            for wallet in target_wallets:
                if not wallet.gl_sub_code:
                    continue
                acc_name = frappe.db.sql(
                    "SELECT name FROM `tabAccount` WHERE account_number=%s", wallet.gl_sub_code)
                if not acc_name:
                    continue
                acc_name = acc_name[0][0]

                je_accounts.append({
                    "account": acc_name,
                    "debit_in_account_currency": 0,
                    "credit_in_account_currency": self.amount,
                    "cost_center": cost_center,
                    "user_remark": f"Allocation to {wallet.branch}"
                })
                total_amount += self.amount

            if not je_accounts:
                frappe.throw(
                    _("No valid Accounts found for the selected branches."))

            # 3. Add Debit Line (Source HO)
            if not self.source_bank_account:
                self.set_default_source_account()
            if not self.source_bank_account:
                frappe.throw(_("Source Bank Account (HO) is missing."))

            je_accounts.append({
                "account": self.source_bank_account,
                "debit_in_account_currency": total_amount,
                "credit_in_account_currency": 0,
                "cost_center": cost_center,
                "user_remark": f"Total Fund Allocation for {len(target_wallets)} branches"
            })

            # 4. Create JE [FIXED METHOD]
            # Use frappe.get_doc() instead of new_doc() when passing a full dict structure
            je = frappe.get_doc({
                "doctype": "Journal Entry",
                "voucher_type": "Journal Entry",
                "posting_date": self.transaction_date,
                "company": company,
                "user_remark": f"Fund Allocation Ref: {self.name}",
                "accounts": je_accounts  # Passing list of dicts works with get_doc
            })

            # Flags
            je.flags.ignore_permissions = True
            je.insert(ignore_permissions=True)

            self.journal_entry_ref = je.name

            # Important: Update the current doc with the ref without saving again (avoid recursion)
            self.db_set("journal_entry_ref", je.name)

            if self.is_bulk_allocation:
                frappe.msgprint(_("Draft Journal Entry created for {0} branches. Total: ₹{1}").format(
                    len(target_wallets), total_amount))

        finally:
            frappe.set_user(original_user)

    def process_finacle_transfer(self):
        if not self.journal_entry_ref:
            frappe.throw(_("Journal Entry Reference is missing."))

        # Call API
        response = individual_finacle_fund_transfer_api(self.journal_entry_ref)
        status = response.get("status")

        if status == "SUCCESS":
            tran_id = response.get("trn_id")
            je_doc = frappe.get_doc("Journal Entry", self.journal_entry_ref)
            if je_doc.docstatus == 0:
                je_doc.flags.ignore_permissions = True
                je_doc.submit()

            self.db_set('finacle_tran_id', tran_id)
            self.db_set('finacle_tran_date', nowdate())
            self.db_set('approval_status', 'Posted')
            # frappe.msgprint(_(f"Finacle Transfer Successful! ID: {tran_id}"), indicator='green')
            frappe.msgprint(
                msg="Finacle Transfer Successful! ID: {tran_id}",
                title="Success",
                indicator='green',
                alert=True  # <--- Added this to make it a floating alert
            )

        else:
            error_msg = response.get("message", "Unknown Finacle Error")
            # Just Throw. Doc stays Draft.
            # BUT 'submission_attempted' will remain 1 because JS set it separately!
            # frappe.throw(_(f"Finacle Transaction Failed: {error_msg}"))
            frappe.msgprint(
                msg="Finacle Transaction Failed: {error_msg}",
                title="Error",
                indicator='red',
                alert=True  # <--- Added this to make it a floating alert
            )

    def create_journal_entry(self):
        """
        Creates a Draft Journal Entry. 
        [UPDATED] Uses custom fields for Remarks and Date to avoid conflicts.
        """
        # 1. Get Credit Account
        wallet = frappe.get_doc("Branch Petty Cash Account", {
                                "branch": self.branch})
        if not wallet.gl_sub_code:
            frappe.throw(_("Branch Wallet has no GL Sub Code defined."))

        credit_account = self.get_or_create_account(
            gl_code=wallet.gl_sub_code,
            account_name=f"Petty Cash - {self.branch_name}",
            parent_group="Sahayog Petty Cash Wallets"
        )

        # Get or Create Cost Center
        valid_cost_center = self.get_or_create_cost_center()

        # 2. Prepare Journal Accounts
        accounts = []
        total_credit = 0.0

        for item in self.items:
            amount = flt(item.amount)
            if amount <= 0:
                continue

            if not item.finacle_gl_code:
                frappe.throw(
                    _("Row #{0}: Missing Finacle GL Code.").format(item.idx))

            # Find/Create Debit Account
            debit_account = self.get_or_create_account(
                gl_code=item.finacle_gl_code,
                account_name=f"{item.expense_category} - {self.branch_name}",
                parent_group="Sahayog Branch Expenses"
            )

            # Add Debit Line
            accounts.append({
                "account": debit_account,
                "debit_in_account_currency": amount,
                "credit_in_account_currency": 0,
                "cost_center": valid_cost_center,
                "user_remark": f"{item.description} (Bill: {item.bill_number})"
            })
            total_credit += amount

        # 3. Add Credit Line (Total)
        accounts.append({
            "account": credit_account,
            "debit_in_account_currency": 0,
            "credit_in_account_currency": total_credit,
            "cost_center": valid_cost_center,
            "user_remark": f"Total Petty Cash Expense for {self.name}"
        })

        # 4. Create Journal Entry Doc
        je = frappe.get_doc({
            "doctype": "Journal Entry",
            "voucher_type": "Journal Entry",
            "posting_date": self.transaction_date,
            "company": frappe.defaults.get_user_default("Company"),
            "accounts": accounts,

            # --- UPDATED FIELDS ---
            "cheque_no": "",
            # "cheque_date": self.transaction_date,  <-- REMOVED (Conflicted)

            # Map to NEW Custom Fields
            "cheque_date": self.transaction_date,
            "cheque_no": f"Petty Cash Expense: {self.name}",

            # Keep standard remark generic or empty to avoid conflict
            # "user_remark": "Auto-generated from Petty Cash",

            "reference_type": "Petty Cash Transaction",
            "reference_name": self.name
        })

        # 5. Save (Status: Draft)
        je.insert(ignore_permissions=True)

        # Link JE back to this doc
        self.db_set("journal_entry_ref", je.name)

        # frappe.msgprint(_("Journal Entry created: {0}").format(je.name))
        frappe.msgprint(
            msg=(_("Journal Entry created: {0}").format(je.name)),
            title="Message",
            indicator='blue',
            alert=True  # <--- Added this to make it a floating alert
        )

    def get_or_create_account(self, gl_code, account_name, parent_group):
        """
        Helper: Checks if Account exists by GL Code. If not, creates it.
        """
        # 1. Define Company
        company = frappe.defaults.get_user_default(
            "Company") or frappe.db.get_single_value("Global Defaults", "default_company")
        if not company:
            frappe.throw("Default Company is not set.")

        # 2. Check if Account already exists
        existing = frappe.db.get_value(
            "Account", {"account_number": gl_code, "company": company}, "name")
        if existing:
            return existing

        # 3. Find Parent Account
        parent_acc_name = frappe.db.get_value(
            "Account", {"account_name": parent_group, "company": company}, "name")
        if not parent_acc_name:
            frappe.throw(
                f"Parent Account Group '{parent_group}' not found for company '{company}'. Please create it in Chart of Accounts.")

        # 4. Verify Parent is a Group
        is_group = frappe.db.get_value("Account", parent_acc_name, "is_group")
        if not is_group:
            frappe.throw(
                f"Account '{parent_acc_name}' exists but is NOT a Group. Please check 'Is Group' in Chart of Accounts.")

        # 5. Determine Correct Account Type
        # If the group has 'Expenses' in the name, we assume it's a Direct Expense
        if "Expenses" in parent_group:
            acc_type = "Direct Expense"
        else:
            acc_type = "Cash"

        # 6. Create the New Account
        new_account = frappe.get_doc({
            "doctype": "Account",
            "account_name": account_name,
            "account_number": gl_code,
            "company": company,
            "parent_account": parent_acc_name,
            "account_type": acc_type,  # <--- FIXED: Now uses "Direct Expense"
            "currency": "INR"
        })
        new_account.insert(ignore_permissions=True)

        return new_account.name

    # new flow

    def on_cancel(self):
        if self.transaction_type == "Expense":
            if self.is_legacy_unsettled_cash_flow_enabled():
                wallet = self.get_branch_wallet_doc()
                wallet.update_unsettled_cash(self.amount, "Withdrawal")
            else:
                self.apply_new_flow_terminal_credit("Canceled")

            self.db_set("approval_status", "Canceled", update_modified=False)

        self.update_wallet()

    def update_wallet(self):
        # Trigger wallet update
        if frappe.flags.in_test:
            return

        # [FIX] We REMOVED the line: wallet.current_balance = wallet.get_current_balance()
        # This prevents the system from overwriting the Finacle Balance with a manual calculation.

        # We only update 'last_funded_on' for Fund Allocations
        if self.transaction_type == "Fund Allocation" and self.docstatus == 1:
            frappe.db.set_value("Branch Petty Cash Account", {
                                "branch": self.branch}, "last_funded_on", self.transaction_date)

    @frappe.whitelist()
    def ho_approve_limit(self):
        user_roles = frappe.get_roles()

        if not set(["HO Petty Cash Manager", "HO Petty Cash Approver"]).intersection(user_roles) and frappe.session.user != "Administrator":
            frappe.throw(
                _("Only HO Petty Cash Manager or Approver can approve."))

        if self.docstatus != 1 or self.approval_status != "Pending Approval":
            frappe.throw(_("Document is not pending approval."))

        self.db_set("amount_deducted", self.amount, update_modified=False)
        self.db_set("approval_status", "Approved", update_modified=False)

        if self.is_legacy_unsettled_cash_flow_enabled():
            self.update_wallet()
        else:
            self.apply_new_flow_pending_approval_deduction()

        frappe.msgprint(
            msg=_("Limit Exceedance Approved. Full amount approved."),
            title=_("Message"),
            indicator="blue"
        )

    def apply_new_flow_pending_approval_deduction(self):
        if self.transaction_type != "Expense":
            return

        if self.is_legacy_unsettled_cash_flow_enabled():
            return

        if self.approval_status != "Approved":
            return

        already_deducted = flt(self.wallet_effect_amount)
        target_total = flt(self.amount_deducted)
        pending_amount = target_total - already_deducted

        if pending_amount <= 0:
            return

        wallet = self.get_branch_wallet_doc()
        wallet.deduct_current_balance(
            amount=pending_amount,
            reference_doctype=self.doctype,
            reference_name=self.name
        )

        self.db_set("wallet_effect_applied", 1, update_modified=False)
        self.db_set("wallet_effect_amount",
                    target_total, update_modified=False)
        self.db_set("wallet_effect_reference_status",
                    self.approval_status, update_modified=False)

    from frappe.utils import now_datetime

    @frappe.whitelist()
    def ho_verify_bill(self):
        user_roles = frappe.get_roles()

        if not set(["HO Petty Cash Manager", "HO Petty Cash Verifier"]).intersection(user_roles) and frappe.session.user != "Administrator":
            frappe.throw(
                _("Only HO Petty Cash Manager or Verifier can verify."))

        if self.approval_status != "Approved":
            frappe.throw(_("Document must be Approved before Verification."))

        enable_integration = cint(
            frappe.db.get_single_value(
                "Sahayog Settings", "enable_finacle_integration") or 0
        )

        if enable_integration:
            if not self.journal_entry_ref:
                frappe.throw(_("Journal Entry Reference is missing."))

            from sahayog.petty_cash_management.api.finacle_integration import individual_finacle_fund_transfer_api

            response = individual_finacle_fund_transfer_api(
                self.journal_entry_ref)

            if response.get("status") != "SUCCESS":
                frappe.msgprint(
                    msg=_("Finacle Failed: {0}").format(
                        response.get("message")),
                    title=_("Error"),
                    indicator="red"
                )
                return

            frappe.db.set_value(
                self.doctype,
                self.name,
                {
                    "finacle_tran_id": response.get("trn_id"),
                    "finacle_tran_date": now_datetime(),
                    "approved_by": frappe.session.user
                },
                update_modified=False
            )
        else:
            self.db_set("approved_by", frappe.session.user,
                        update_modified=False)

        if self.is_legacy_unsettled_cash_flow_enabled():
            self.db_set("approval_status", "Verified", update_modified=False)
        else:
            self.apply_new_flow_terminal_credit("Verified")
            self.db_set("approval_status", "Verified", update_modified=False)
            self.db_set("finacle_tran_date", now_datetime(),
                        update_modified=False)

        frappe.msgprint(
            msg=_("Bills Verified successfully."),
            title=_("Verification Complete"),
            indicator="green"
        )

    def has_permission(self, ptype="read", user=None):
        if not user:
            user = frappe.session.user

        user_roles = frappe.get_roles(user)

        # [UPDATED] Allow Administrator and all HO roles to open the document
        if user == 'Administrator' or any(r in user_roles for r in ["HO Petty Cash Manager", "HO Petty Cash Approver", "HO Petty Cash Verifier", "System Manager", "HO Petty Cash Auditor",]):
            return True

        # Existing branch logic fallback
        from sahayog.petty_cash_management.permissions import get_user_allowed_branches
        allowed_branches = get_user_allowed_branches(user)

        if allowed_branches is None:  # Admin/Manager fallback
            return True

        if self.is_new():
            return True

        if self.branch in (allowed_branches or []):
            return True

        return False

    def get_or_create_cost_center(self):
        """ 
        [NEW] Automates Cost Center Creation.
        Checks if a Cost Center exists for this Branch (e.g. '1113').
        If not, creates it automatically under a 'Sahayog Branches' group.
        """
        company = frappe.defaults.get_user_default(
            "Company") or frappe.db.get_single_value("Global Defaults", "default_company")

        # 1. Search by Cost Center Number (Branch ID)
        existing_cc = frappe.db.get_value(
            "Cost Center", {"cost_center_number": self.branch, "company": company}, "name")
        if existing_cc:
            return existing_cc

        # 2. If missing, find a Parent Group to attach to
        # Try to find 'Sahayog Branches' or create it
        parent_group_name = "Sahayog Branches"
        parent_cc = frappe.db.get_value(
            "Cost Center", {"cost_center_name": parent_group_name, "company": company}, "name")

        if not parent_cc:
            # Create the Group Node if it doesn't exist
            root_cc_name = frappe.db.get_value("Cost Center", {
                                               # Main Root
                                               "is_group": 1, "company": company, "parent_cost_center": ["is", "not set"]}, "name")

            new_group = frappe.get_doc({
                "doctype": "Cost Center",
                "cost_center_name": parent_group_name,
                "is_group": 1,
                "company": company,
                "parent_cost_center": root_cc_name or "Main - S"  # Fallback
            })
            new_group.insert(ignore_permissions=True)
            parent_cc = new_group.name

        # 3. Create the Branch Cost Center
        new_cc = frappe.get_doc({
            "doctype": "Cost Center",
            "cost_center_name": self.branch_name,  # e.g. "CHEMBUR BRANCH"
            "cost_center_number": self.branch,    # e.g. "1113"
            "company": company,
            "parent_cost_center": parent_cc
        })
        new_cc.insert(ignore_permissions=True)

        return new_cc.name

    def download_transaction_excel(self):
        """
        Generates a detailed CSV report for this specific transaction.
        Includes resolved names for Categories and GL Codes.
        """
        import csv
        import io

        # 1. Fetch Related Data (Optimized)
        # Fetch Wallet GL Code (Single fetch for the parent)
        wallet_gl_code = frappe.db.get_value("Branch Petty Cash Account", {
                                             "branch": self.branch}, "gl_sub_code") or ""

        # Fetch Category Names Map (Fetch all used categories in one query)
        # This prevents N+1 queries inside the loop
        category_ids = [
            row.expense_category for row in self.items if row.expense_category]
        category_map = {}
        if category_ids:
            # Fetch name where name in list
            categories = frappe.get_all("Expense Category", filters={
                                        "name": ["in", category_ids]}, fields=["name", "category_name"])
            for cat in categories:
                category_map[cat.name] = cat.category_name

        # 2. Define CSV Headers
        headers = [
            "Transaction ID", "Branch Code", "Branch Name", "Date", "Type",
            "Total Amount", "Wallet Balance", "Cash in Hand",
            "Approval Status", "Within Limit", "Exceeding Limit", "Deducted Amount",
            "Approved By", "TTUM Remarks", "Finacle Remarks", "Branch Petty Cash Account",
            # Child Item Fields
            "Expense Category", "Vendor", "Bill No", "Item Amount", "Description", "Expense GL Code"
        ]

        # 3. Prepare Data Rows
        rows = []

        # Helper to format amounts safely
        def fmt(val):
            return flt(val) if val else 0.0

        for row in self.items:
            # Resolve Category Name
            cat_name = category_map.get(
                row.expense_category, row.expense_category)

            rows.append([
                self.name,
                self.branch,
                self.branch_name,
                self.transaction_date,
                self.transaction_type,
                fmt(self.amount),
                fmt(self.current_branch_balance),
                fmt(self.current_unsettled_cash),
                self.approval_status,
                fmt(self.amount_within_limit),
                fmt(self.amount_exceeding_limit),
                fmt(self.amount_deducted),
                self.approved_by,
                self.custom_ttum_remarks,
                self.finacle_tran_particular,
                wallet_gl_code,
                # Child Data
                cat_name,
                row.vendor_name,
                row.bill_number,
                fmt(row.amount),
                row.description,
                row.finacle_gl_code
            ])

        # 4. Generate CSV using Python's built-in csv module
        # We use StringIO to write to memory first
        output = io.StringIO()
        writer = csv.writer(output)

        # Write BOM for Excel compatibility (optional but recommended for UTF-8)
        output.write('\ufeff')

        writer.writerow(headers)
        writer.writerows(rows)

        # 5. Set Response
        output.seek(0)
        csv_content = output.getvalue()
        output.close()

        frappe.response['filename'] = f"{self.name}.csv"
        frappe.response['filecontent'] = csv_content

        # [FIX] Use 'binary' instead of 'csv' to bypass Frappe's auto-CSV logic
        # We manually created the CSV content, so we treat it as a file download.
        frappe.response['type'] = 'binary'

    def download_transaction_txt(self):
        content = []
        from frappe.utils import getdate
        date_obj = getdate(self.transaction_date)
        ttum_date = date_obj.strftime("%b%y").upper()  # JAN26
        currency_str = f"INR{self.branch}"

        narrative_suffix = self.custom_ttum_remarks if self.custom_ttum_remarks else f"{ttum_date} {self.name}"
        debitDescription = ""
        total_debit = 0.0

        # --- 1. DEBIT ROWS (Expenses) ---
        for row in self.items:
            if not row.finacle_gl_code:
                frappe.throw(f"Row #{row.idx} is missing Finacle GL Code")

            amount_str = "{:.2f}".format(row.amount)
            total_debit += row.amount

            # Generate the description and restrict it to exactly 30 characters
            raw_desc = f"{row.description}" if row.description else narrative_suffix
            debitDescription = raw_desc[:30]  # <--- ADDED 30 CHAR LIMIT HERE

            # --- SPACING LOGIC ---
            # Standard Finacle width is 17.
            # Logic: Calculate space for 17 width. If < 10, force 10.
            padding_count = 17 - len(amount_str)
            if padding_count < 10:
                padding_count = 10

            space_str = " " * padding_count
            # ---------------------

            # Format: GL <1sp> CURR <4sp> D <padding> AMOUNT REMARKS
            line = f"{row.finacle_gl_code} {currency_str}    D{space_str}{amount_str}{debitDescription}"
            content.append(line)

        # --- 2. CREDIT ROW (Branch Wallet) ---
        wallet_gl = frappe.db.get_value("Branch Petty Cash Account", {
                                        "branch": self.branch}, "gl_sub_code")
        if not wallet_gl:
            frappe.throw(f"GL Sub Code not found for Branch {self.branch}")

        total_amount_str = "{:.2f}".format(total_debit)

        # --- SPACING LOGIC (Same for Credit) ---
        padding_count = 17 - len(total_amount_str)
        if padding_count < 10:
            padding_count = 10

        space_str = " " * padding_count
        # ---------------------------------------

        # Restrict the credit narrative to exactly 30 characters
        # <--- ADDED 30 CHAR LIMIT HERE
        creditDescription = narrative_suffix[:30]

        credit_line = f"{wallet_gl} {currency_str}    C{space_str}{total_amount_str}{creditDescription}"
        content.append(credit_line)

        final_txt = "\n".join(content)

        frappe.response['filename'] = f"TTUM_{self.branch}_{self.name}.txt"
        frappe.response['filecontent'] = final_txt
        frappe.response['type'] = 'download'

    def validate_item_descriptions(self):
        for row in self.items:
            if row.description and len(row.description) > 30:
                frappe.throw(
                    f"Row #{row.idx}: Description cannot be more than 30 characters including spaces."
                )

    def is_legacy_unsettled_cash_flow_enabled(self):
        return cint(
            frappe.db.get_single_value(
                "Sahayog Settings", "enable_unsettled_cash_flow") or 0
        )

    def get_branch_wallet_doc(self):
        wallet_name = frappe.db.get_value("Branch Petty Cash Account", {
                                          "branch": self.branch}, "name")
        if not wallet_name:
            frappe.throw(
                _("Branch Petty Cash Account for branch {0} not found.").format(self.branch))
        return frappe.get_doc("Branch Petty Cash Account", wallet_name)

    # def get_effective_deduction_amount(self):
    #     return flt(self.amount_deducted) if flt(self.amount_deducted) > 0 else flt(self.amount)

    def get_effective_deduction_amount(self):
        return flt(self.amount_deducted)

    def apply_new_flow_wallet_deduction(self):
        if self.transaction_type != "Expense":
            return

        if self.is_legacy_unsettled_cash_flow_enabled():
            return

        if cint(self.wallet_effect_applied):
            return

        if self.approval_status not in ["Approved", "Pending Approval"]:
            return

        deduct_amount = flt(self.get_effective_deduction_amount())

        if deduct_amount <= 0:
            self.db_set("wallet_effect_applied", 0, update_modified=False)
            self.db_set("wallet_effect_amount", 0, update_modified=False)
            self.db_set("wallet_effect_reference_status",
                        self.approval_status, update_modified=False)
            return

        wallet = self.get_branch_wallet_doc()
        wallet.deduct_current_balance(
            amount=deduct_amount,
            reference_doctype=self.doctype,
            reference_name=self.name
        )

        self.db_set("wallet_effect_applied", 1, update_modified=False)
        self.db_set("wallet_effect_amount",
                    deduct_amount, update_modified=False)
        self.db_set("wallet_effect_reference_status",
                    self.approval_status, update_modified=False)

    def apply_new_flow_terminal_credit(self, action):
        if self.transaction_type != "Expense":
            return

        if self.is_legacy_unsettled_cash_flow_enabled():
            return

        if not cint(self.wallet_effect_applied):
            return

        if self.wallet_terminal_action:
            return

        credit_amount = flt(self.wallet_effect_amount)
        if credit_amount <= 0:
            return

        wallet = self.get_branch_wallet_doc()
        wallet.credit_current_balance(
            amount=credit_amount,
            reference_doctype=self.doctype,
            reference_name=self.name
        )

        self.db_set("wallet_terminal_action", action, update_modified=False)

    def validate_expense_against_active_wallet(self):
        if self.transaction_type != "Expense":
            return

        wallet_data = frappe.db.get_value(
            "Branch Petty Cash Account",
            {"branch": self.branch},
            ["current_balance", "unsettled_cash"],
            as_dict=True
        )

        if not wallet_data:
            frappe.throw(
                _("Branch Petty Cash Account not found for branch {0}.").format(self.branch))

        current_balance = flt(wallet_data.current_balance)
        unsettled_cash = flt(wallet_data.unsettled_cash)

        if self.is_legacy_unsettled_cash_flow_enabled():
            total_buying_power = current_balance + unsettled_cash
            if total_buying_power < flt(self.amount):
                frappe.throw(
                    _("Insufficient Funds. Bank: {0}, Cash-in-Hand: {1}, Total: {2}, Required: {3}").format(
                        current_balance, unsettled_cash, total_buying_power, self.amount
                    )
                )
        else:
            if current_balance <= 0:
                frappe.throw(
                    _("Current Balance is 0 for branch {0}. You cannot submit an expense transaction.").format(
                        self.branch
                    )
                )

            if flt(self.amount) > current_balance:
                frappe.throw(
                    _("You cannot submit this expense because the available Current Balance for branch {0} is {1} and requested expense amount is {2}. Current Balance can never go negative.").format(
                        self.branch, current_balance, self.amount
                    )
                )


@frappe.whitelist()
def get_category_limit_status(branch, category, transaction_date, docname=None):
    """
    API used by Client Script to show available limit.
    """
    if not branch or not category or not transaction_date:
        return 0

    branch_type = frappe.db.get_value("Branch Petty Cash Account", {
                                      "branch": branch}, "branch_type")
    category_doc = frappe.get_doc("Expense Category", category)
    limit = category_doc.metro_limit if branch_type == "Metro" else category_doc.non_metro_limit

    if limit == 0:
        return 999999999

    first_day = get_first_day(transaction_date)
    last_day = get_last_day(transaction_date)

    # Same SQL Update Here
    spent_sql = """
            SELECT COALESCE(SUM(child.amount), 0)
            FROM `tabPetty Cash Transaction Item` child
            JOIN `tabPetty Cash Transaction` parent ON child.parent = parent.name
            WHERE parent.branch = %s 
              AND child.expense_category = %s
              AND parent.transaction_date BETWEEN %s AND %s
              AND parent.docstatus = 1
              AND parent.approval_status != 'Verified'
              AND parent.name != %s
        """

    already_spent = frappe.db.sql(
        spent_sql, (branch, category, first_day, last_day, docname or "New"))[0][0]

    return max(flt(limit) - flt(already_spent), 0)


# new flow

@frappe.whitelist()
def get_branch_balance(branch):
    if not branch:
        return

    data = frappe.db.get_value(
        "Branch Petty Cash Account",
        {"branch": branch},
        ["current_balance", "unsettled_cash"],
        as_dict=True
    )

    return {
        "current_balance": flt(data.current_balance) if data else 0.0,
        "unsettled_cash": flt(data.unsettled_cash) if data else 0.0
    }


@frappe.whitelist()
def get_ho_source_account():
    """
    Returns the Account Name (ID) for the branch marked as 'Is Fund Source'.
    Used by Client Script to auto-populate the field.
    """
    # 1. Find the Wallet
    ho_wallet = frappe.db.get_value("Branch Petty Cash Account",
                                    {"is_fund_source": 1, "status": "Active"},
                                    ["gl_sub_code"],
                                    as_dict=True
                                    )

    if not ho_wallet or not ho_wallet.gl_sub_code:
        return None

    # 2. Find the Account
    # Use SQL for speed/permission bypass
    account = frappe.db.sql(
        """SELECT name FROM `tabAccount` WHERE account_number=%s""", ho_wallet.gl_sub_code)

    return account[0][0] if account else None


@frappe.whitelist()
def mark_submission_attempt(docname):
    # Use update_modified=False to prevent timestamp change
    frappe.db.set_value("Petty Cash Transaction", docname,
                        "submission_attempted", 1, update_modified=False)
    frappe.db.commit()


@frappe.whitelist()
def download_transaction_report(filters=None):
    """
    Generate Excel report for Petty Cash Transactions with applied filters.
    Uses frappe.get_all to automatically handle List View filters.
    """
    import json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from frappe.utils.file_manager import save_file
    import tempfile
    import os
    import base64
    from datetime import datetime

    # 1. Parse filters if passed as JSON string
    if filters and isinstance(filters, str):
        filters = json.loads(filters)

    # 2. Fetch Data using frappe.get_all (Handles standard List View filters automatically)
    # We explicitly list the fields to ensure we get the data we need
    fields = [
        "name",
        "transaction_type",
        "transaction_date",
        "branch",
        "branch_name",
        "amount",
        "current_branch_balance",
        "current_unsettled_cash",
        "approval_status",
        "amount_within_limit",
        "amount_exceeding_limit",
        "amount_deducted",
        "approved_by",
        "finacle_tran_id",
        "finacle_tran_date",
        "remarks",
        "docstatus",
        "creation",
        "modified",
        "owner",
        "modified_by"
    ]

    # Note: frappe.get_all automatically respects permissions and filters
    transactions = frappe.get_all(
        "Petty Cash Transaction",
        filters=filters,
        fields=fields,
        order_by="transaction_date desc, creation desc"
    )

    if not transactions:
        frappe.throw(_("No records found matching the filters."))

    # 3. Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Petty Cash Transactions"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Transaction ID", "Transaction Type", "Date", "Branch Code", "Branch Name",
        "Amount (₹)", "Balance (₹)", "Cash in Hand (₹)", "Approval Status",
        "Within Limit (₹)", "Exceeding Limit (₹)", "Amount Deducted (₹)",
        "Approved By", "Finacle Txn ID", "Finacle Date", "Remarks",
        "Status", "Created On", "Modified On", "Created By", "Modified By"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 4. Write data
    for row_num, txn in enumerate(transactions, 2):
        # Map docstatus to readable text
        status_map = {0: "Draft", 1: "Submitted", 2: "Cancelled"}

        row_data = [
            txn.get('name'),
            txn.get('transaction_type'),
            txn.get('transaction_date').strftime(
                '%d-%m-%Y') if txn.get('transaction_date') else '',
            txn.get('branch'),
            txn.get('branch_name'),
            txn.get('amount', 0),
            txn.get('current_branch_balance', 0),
            txn.get('current_unsettled_cash', 0),
            txn.get('approval_status'),
            txn.get('amount_within_limit', 0),
            txn.get('amount_exceeding_limit', 0),
            txn.get('amount_deducted', 0),
            txn.get('approved_by'),
            txn.get('finacle_tran_id'),
            txn.get('finacle_tran_date').strftime(
                '%d-%m-%Y') if txn.get('finacle_tran_date') else '',
            txn.get('remarks'),
            status_map.get(txn.get('docstatus'), 'Unknown'),
            txn.get('creation').strftime(
                '%d-%m-%Y %H:%M') if txn.get('creation') else '',
            txn.get('modified').strftime(
                '%d-%m-%Y %H:%M') if txn.get('modified') else '',
            txn.get('owner'),
            txn.get('modified_by')
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            # Format currency columns
            if col_num in [6, 7, 8, 10, 11, 12]:  # Amount columns
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # 5. Handle File Output
    # Create temp file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    # Read file content
    with open(temp_file.name, 'rb') as f:
        content = f.read()

    # Delete temp file
    os.unlink(temp_file.name)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Petty_Cash_Transactions_{timestamp}.xlsx"

    # Return as base64 encoded string
    file_data = base64.b64encode(content).decode('utf-8')

    return {
        'filename': filename,
        'filecontent': file_data,
        'record_count': len(transactions)
    }


@frappe.whitelist()
def download_excel_api(name):   # <--- Renamed
    doc = frappe.get_doc("Petty Cash Transaction", name)
    doc.download_transaction_excel()


@frappe.whitelist()
def download_txt_api(name):     # <--- Renamed
    doc = frappe.get_doc("Petty Cash Transaction", name)
    doc.download_transaction_txt()


@frappe.whitelist()
def download_consolidated_txt_api(transaction_date=None):
    """Generates a Consolidated TTUM Text file for all Verified transactions of the selected date."""
    from frappe.utils import nowdate, getdate

    user_roles = frappe.get_roles()
    if frappe.session.user != "Administrator" and not set(
        ["HO Petty Cash Manager", "HO Petty Cash Verifier"]
    ).intersection(user_roles):
        frappe.throw(
            "Only HO Petty Cash Manager, Verifier, or Administrator can download consolidated files."
        )

    selected_date = str(getdate(transaction_date)
                        ) if transaction_date else nowdate()

    transactions = frappe.get_all(
        "Petty Cash Transaction",
        filters={
            "finacle_tran_date": selected_date,
            "approval_status": "Verified"
        },
        pluck="name",
        order_by="creation ASC"
    )

    if not transactions:
        return {
            "status": "no_data",
            "message": f"No verified transactions found for {selected_date}."
        }

    content = []

    for txn_name in transactions:
        doc = frappe.get_doc("Petty Cash Transaction", txn_name)

        date_obj = getdate(doc.finacle_tran_date)
        ttum_date = date_obj.strftime("%b%y").upper()
        currency_str = f"INR{doc.branch}"

        narrative_suffix = (
            doc.custom_ttum_remarks
            if doc.custom_ttum_remarks
            else f"{ttum_date} {doc.name}"
        )

        total_debit = 0.0

        for row in doc.items:
            if not row.finacle_gl_code:
                frappe.throw(
                    f"Row #{row.idx} in {doc.name} is missing Finacle GL Code")

            amount_str = "{:.2f}".format(row.amount)
            total_debit += row.amount

            raw_desc = row.description if row.description else narrative_suffix
            debit_description = raw_desc[:30]

            padding_count = 17 - len(amount_str)
            if padding_count < 10:
                padding_count = 10
            space_str = " " * padding_count

            line = f"{row.finacle_gl_code} {currency_str}    D{space_str}{amount_str}{debit_description}"
            content.append(line)

        wallet_gl = frappe.db.get_value(
            "Branch Petty Cash Account",
            {"branch": doc.branch},
            "gl_sub_code"
        )
        if not wallet_gl:
            frappe.throw(f"GL Sub Code not found for Branch {doc.branch}")

        total_amount_str = "{:.2f}".format(total_debit)
        padding_count = 17 - len(total_amount_str)
        if padding_count < 10:
            padding_count = 10
        space_str = " " * padding_count
        credit_description = narrative_suffix[:30]

        credit_line = f"{wallet_gl} {currency_str}    C{space_str}{total_amount_str}{credit_description}"
        content.append(credit_line)

    final_txt = "\n".join(content)

    return {
        "status": "success",
        "filename": f"Consolidated_TTUM_{selected_date}.txt",
        "filecontent": final_txt
    }


@frappe.whitelist()
def download_consolidated_excel_api(transaction_date=None):
    """Generates a Consolidated CSV/Excel report for all Verified transactions of the selected date."""
    from frappe.utils import nowdate, getdate
    import csv
    import io
    import base64

    # 1. Permission Check
    user_roles = frappe.get_roles()
    if frappe.session.user != "Administrator" and not set(
        ["HO Petty Cash Manager", "HO Petty Cash Verifier"]
    ).intersection(user_roles):
        frappe.throw(
            "Only HO Petty Cash Manager, Verifier, or Administrator can download consolidated files."
        )

    selected_date = str(getdate(transaction_date)
                        ) if transaction_date else nowdate()

    # 2. Fetch all matching transactions (Selected Date + Verified)
    transactions = frappe.get_all(
        "Petty Cash Transaction",
        filters={
            "finacle_tran_date": selected_date,
            "approval_status": "Verified"
        },
        pluck="name",
        order_by="creation ASC"
    )

    if not transactions:
        return {
            "status": "no_data",
            "message": f"No verified transactions found for {selected_date}."
        }

    # 3. Prepare CSV content in memory
    output = io.StringIO()
    output.write("\ufeff")  # BOM for Excel UTF-8 compatibility
    writer = csv.writer(output)

    headers = [
        "Transaction ID",
        "Branch Code",
        "Branch Name",
        "Date",
        "Type",
        "Total Amount",
        "Wallet Balance",
        "Cash in Hand",
        "Approval Status",
        "Within Limit",
        "Exceeding Limit",
        "Deducted Amount",
        "Approved By",
        "TTUM Remarks",
        "Finacle Remarks",
        "Branch Petty Cash Account",
        "Expense Category",
        "Vendor Name",
        "Bill Number",
        "Item Amount",
        "Description",
        "Finacle GL Code"
    ]
    writer.writerow(headers)

    # 4. Loop through each transaction and expand child items into separate rows
    for txn_name in transactions:
        doc = frappe.get_doc("Petty Cash Transaction", txn_name)

        wallet_gl_code = frappe.db.get_value(
            "Branch Petty Cash Account",
            {"branch": doc.branch},
            "gl_sub_code"
        ) or ""

        category_ids = [
            row.expense_category for row in doc.items if row.expense_category]
        category_map = {}

        if category_ids:
            categories = frappe.get_all(
                "Expense Category",
                filters={"name": ["in", category_ids]},
                fields=["name", "category_name"]
            )
            category_map = {cat.name: cat.category_name for cat in categories}

        for row in doc.items:
            cat_name = category_map.get(
                row.expense_category, row.expense_category)

            row_data = [
                doc.name,
                doc.branch,
                doc.branch_name,
                doc.transaction_date,
                doc.transaction_type,
                "{:.2f}".format(doc.amount or 0),
                "{:.2f}".format(doc.current_branch_balance or 0),
                "{:.2f}".format(doc.current_unsettled_cash or 0),
                doc.approval_status,
                "{:.2f}".format(doc.amount_within_limit or 0),
                "{:.2f}".format(doc.amount_exceeding_limit or 0),
                "{:.2f}".format(doc.amount_deducted or 0),
                doc.approved_by,
                doc.custom_ttum_remarks,
                doc.finacle_tran_particular,
                wallet_gl_code,
                cat_name,
                row.vendor_name,
                row.bill_number,
                "{:.2f}".format(row.amount or 0),
                row.description,
                row.finacle_gl_code
            ]
            writer.writerow(row_data)

    csv_content = output.getvalue()
    output.close()

    return {
        "status": "success",
        "filename": f"Consolidated_Report_{selected_date}.csv",
        "filecontent": base64.b64encode(csv_content.encode("utf-8")).decode("utf-8")
    }


@frappe.whitelist()
def download_detailed_report_by_date_range(from_date=None, to_date=None):
    """
    Download detailed Excel report for all Petty Cash Transactions
    between from_date and to_date based on transaction_date.
    Each child item is expanded as a separate row.
    """
    import base64
    import tempfile
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from frappe.utils import getdate, flt
    from datetime import datetime

    if not from_date or not to_date:
        frappe.throw("Start Date and End Date are required.")

    from_date = getdate(from_date)
    to_date = getdate(to_date)

    if from_date > to_date:
        frappe.throw("Start Date cannot be greater than End Date.")

    transaction_names = frappe.get_all(
        "Petty Cash Transaction",
        filters=[
            ["transaction_date", ">=", from_date],
            ["transaction_date", "<=", to_date]
        ],
        pluck="name",
        order_by="transaction_date asc, creation asc"
    )

    if not transaction_names:
        frappe.throw(f"No records found between {from_date} and {to_date}.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Petty Cash Report"

    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "Transaction ID", "Branch Code", "Branch Name", "Date", "Type",
        "Total Amount", "Wallet Balance", "Cash in Hand",
        "Approval Status", "Within Limit", "Exceeding Limit", "Deducted Amount",
        "Approved By", "TTUM Remarks", "Finacle Remarks", "Branch Petty Cash Account",
        "Expense Category", "Vendor", "Bill No", "Bill Date", "Item Amount",
        "Description", "Expense GL Code", "Bill Attachment",
        "Posted To Finacle", "Finacle Txn ID", "Finacle Date",
        "Remarks", "Doc Status", "Created On", "Modified On", "Created By", "Modified By"
    ]

    for colnum, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=colnum)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def fmt(val):
        return flt(val) if val else 0.0

    excel_row = 2
    status_map = {0: "Draft", 1: "Submitted", 2: "Cancelled"}

    for txn_name in transaction_names:
        doc = frappe.get_doc("Petty Cash Transaction", txn_name)

        wallet_gl_code = frappe.db.get_value(
            "Branch Petty Cash Account",
            {"branch": doc.branch},
            "gl_sub_code"
        ) or ""

        category_ids = [
            row.expense_category for row in doc.items if row.expense_category]
        category_map = {}

        if category_ids:
            categories = frappe.get_all(
                "Expense Category",
                filters={"name": ["in", category_ids]},
                fields=["name", "category_name"]
            )
            for cat in categories:
                category_map[cat.name] = cat.category_name

        if doc.items:
            for row in doc.items:
                cat_name = category_map.get(
                    row.expense_category, row.expense_category)

                row_data = [
                    doc.name,
                    doc.branch,
                    doc.branch_name,
                    doc.transaction_date.strftime(
                        "%d-%m-%Y") if doc.transaction_date else "",
                    doc.transaction_type,
                    fmt(doc.amount),
                    fmt(doc.current_branch_balance),
                    fmt(doc.current_unsettled_cash),
                    doc.approval_status,
                    fmt(doc.amount_within_limit),
                    fmt(doc.amount_exceeding_limit),
                    fmt(doc.amount_deducted),
                    doc.approved_by,
                    doc.custom_ttum_remarks,
                    doc.finacle_tran_particular,
                    wallet_gl_code,
                    cat_name,
                    row.vendor_name,
                    row.bill_number,
                    row.bill_date.strftime(
                        "%d-%m-%Y") if row.bill_date else "",
                    fmt(row.amount),
                    row.description,
                    row.finacle_gl_code,
                    row.bill_attachment,
                    doc.posted_to_finacle,
                    doc.finacle_tran_id,
                    doc.finacle_tran_date.strftime(
                        "%d-%m-%Y") if doc.finacle_tran_date else "",
                    doc.remarks,
                    status_map.get(doc.docstatus, "Unknown"),
                    doc.creation.strftime(
                        "%d-%m-%Y %H:%M") if doc.creation else "",
                    doc.modified.strftime(
                        "%d-%m-%Y %H:%M") if doc.modified else "",
                    doc.owner,
                    doc.modified_by
                ]

                for colnum, value in enumerate(row_data, 1):
                    cell = ws.cell(row=excel_row, column=colnum)
                    cell.value = value
                    cell.border = border

                    if colnum in [6, 7, 8, 10, 11, 12, 21]:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(vertical="top")

                excel_row += 1
        else:
            row_data = [
                doc.name,
                doc.branch,
                doc.branch_name,
                doc.transaction_date.strftime(
                    "%d-%m-%Y") if doc.transaction_date else "",
                doc.transaction_type,
                fmt(doc.amount),
                fmt(doc.current_branch_balance),
                fmt(doc.current_unsettled_cash),
                doc.approval_status,
                fmt(doc.amount_within_limit),
                fmt(doc.amount_exceeding_limit),
                fmt(doc.amount_deducted),
                doc.approved_by,
                doc.custom_ttum_remarks,
                doc.finacle_tran_particular,
                wallet_gl_code,
                "", "", "", "", 0.0, "", "", "",
                doc.posted_to_finacle,
                doc.finacle_tran_id,
                doc.finacle_tran_date.strftime(
                    "%d-%m-%Y") if doc.finacle_tran_date else "",
                doc.remarks,
                status_map.get(doc.docstatus, "Unknown"),
                doc.creation.strftime(
                    "%d-%m-%Y %H:%M") if doc.creation else "",
                doc.modified.strftime(
                    "%d-%m-%Y %H:%M") if doc.modified else "",
                doc.owner,
                doc.modified_by
            ]

            for colnum, value in enumerate(row_data, 1):
                cell = ws.cell(row=excel_row, column=colnum)
                cell.value = value
                cell.border = border
            excel_row += 1

    for colnum in range(1, len(headers) + 1):
        column_letter = get_column_letter(colnum)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    ws.freeze_panes = "A2"

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    with open(tmp.name, "rb") as f:
        content = f.read()

    os.unlink(tmp.name)

    filedata = base64.b64encode(content).decode("utf-8")
    filename = f"Petty_Cash_Detailed_Report_{from_date}_to_{to_date}.xlsx"

    return {
        "filename": filename,
        "filecontent": filedata,
        "recordcount": len(transaction_names)
    }
